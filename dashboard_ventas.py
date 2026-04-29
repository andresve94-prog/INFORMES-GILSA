"""
Dashboard de Ventas
Ejecutar con:  streamlit run dashboard_ventas.py
"""

import re
import io
import json
import base64
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─── Rutas de archivos ───────────────────────────────────────────────────────

DATA_FILE   = Path(__file__).parent / "VENTAS_GILSA_CONSOLIDADO.xlsx"
CONFIG_FILE = Path(__file__).parent / "config.json"

# ─── Configuracion por defecto ────────────────────────────────────────────────

DEFAULT_CFG = {
    "empresa_nombre": "GILSA S.A.S.",
    "empresa_nit":    "900.007.450-8",
    "logo_b64":       "",
    "logo_mime":      "image/png",
    "color_v1":       "#2563EB",   # TOCARO
    "color_v2":       "#16A34A",   # JUAN GABRIEL
    "color_v3":       "#D97706",   # SIN ASIGNAR
    "nombre_v1":      "TOCARO",
    "nombre_v2":      "JUAN GABRIEL",
    "nombre_v3":      "SIN ASIGNAR",
}

ORDEN_MESES = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE",
]
NUM_A_MES = {i+1: m for i, m in enumerate(ORDEN_MESES)}
MES_A_NUM = {m: i+1 for i, m in enumerate(ORDEN_MESES)}

COLS_DISPLAY = [
    "AÑO","MES","FUENTE","COMPROBANTE","FECHA","CLIENTE",
    "VENTA","DESCUENTO","VENTA NETA","COSTO","UTILIDAD","% UTILIDAD","VENDEDOR",
]

# Valores internos fijos en la base de datos (no cambian con la config)
_V1 = "TOCARO"
_V2 = "JUAN GABRIEL"
_V3 = "SIN ASIGNAR"

# ─── Helpers de configuracion ─────────────────────────────────────────────────

def cfg() -> dict:
    return st.session_state["cfg"]

def get_nombre_map() -> dict:
    c = cfg()
    return {_V1: c["nombre_v1"], _V2: c["nombre_v2"], _V3: c["nombre_v3"]}

def get_colores() -> dict:
    c = cfg()
    return {c["nombre_v1"]: c["color_v1"],
            c["nombre_v2"]: c["color_v2"],
            c["nombre_v3"]: c["color_v3"]}

def con_nombres(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega columna VENDEDOR_DISP con los nombres configurados."""
    df = df.copy()
    df["VENDEDOR_DISP"] = df["VENDEDOR"].map(get_nombre_map()).fillna(df["VENDEDOR"])
    return df

# ─── Page config ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Dashboard de Ventas",
    page_icon=":bar_chart:",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Parser TXT ───────────────────────────────────────────────────────────────

def parse_num(s: str) -> float:
    try:
        return float(s.strip().replace(",", ""))
    except ValueError:
        return 0.0


def find_col_positions(lines):
    for line in lines:
        l = line.rstrip("\n\r")
        if ("VENTA" in l and "COSTO" in l and "TOCARO" in l
                and "JUAN" in l and not l.lstrip().startswith("0")
                and len(l) > 60):
            return {
                "VENTA":     l.index("VENTA"),
                "DESCUENTO": l.index("DESCUENTO") if "DESCUENTO" in l else l.index("VENTA") + 20,
                "COSTO":     l.index("COSTO"),
                "TOCARO":    l.index("TOCARO"),
                "JUAN":      l.index("JUAN"),
            }
    return None


def assign_cols(line: str, col_pos: dict) -> dict:
    cols_sorted = sorted(col_pos.items(), key=lambda x: x[1])
    result = {c: 0.0 for c in col_pos}
    for m in re.finditer(r"[\d,]+\.\d{2}", line):
        num_end = m.end()
        value   = parse_num(m.group())
        best_col, best_start = None, -1
        for name, start in cols_sorted:
            if start <= num_end and start > best_start:
                best_start, best_col = start, name
        if best_col and result[best_col] == 0.0:
            result[best_col] = value
    return result


def detectar_metadata(lines) -> dict | None:
    for line in lines[:20]:
        l = line.strip()
        date_m   = re.search(r"A\s*:\s*(\d{1,2})/(\d{2})/(\d{2})", l)
        fuente_m = re.search(r"FUENTE\s+(\d{2})", l)
        if date_m and fuente_m:
            _, month_str, year_2d = date_m.groups()
            return {"year": 2000 + int(year_2d),
                    "mes_num": int(month_str),
                    "fuente":  fuente_m.group(1)}
    year, mes_num, fuente = None, None, None
    for line in lines[:20]:
        l = line.strip()
        dm = re.search(r"A\s*:\s*\d{1,2}/(\d{2})/(\d{2})", l)
        if dm:
            mes_num = int(dm.group(1))
            year    = 2000 + int(dm.group(2))
        fm = re.search(r"FUENTE\s+(\d{2})", l)
        if fm:
            fuente = fm.group(1)
    if year and mes_num and fuente:
        return {"year": year, "mes_num": mes_num, "fuente": fuente}
    return None


def parsear_contenido(lines, year: int, mes_num: int, fuente: str) -> list[dict]:
    records = []
    col_pos = find_col_positions(lines)
    if not col_pos:
        return records
    venta_start = col_pos["VENTA"]
    for line in lines:
        l = line.rstrip("\n\r")
        m = re.match(r"^(\d{8})\s+(\d{1,2}/\d{2}/\d{2})", l)
        if not m:
            continue
        cpbte   = m.group(1)
        fecha   = m.group(2)
        cliente = re.sub(r"\s+", " ", l[m.end():venta_start]).strip()
        vals    = assign_cols(l, col_pos)
        venta   = vals["VENTA"]
        desc    = vals["DESCUENTO"]
        costo   = vals["COSTO"]
        tocaro  = vals["TOCARO"]
        juan    = vals["JUAN"]
        if venta == 0:
            continue
        if tocaro > 0:
            vendedor = _V1
        elif juan > 0:
            vendedor = _V2
        else:
            vendedor = _V3
        venta_neta = round(venta - desc, 2)
        utilidad   = round(venta_neta - costo, 2)
        pct_util   = round(utilidad / venta_neta, 4) if venta_neta else 0.0
        records.append({
            "AÑO": year, "MES_NUM": mes_num, "MES": NUM_A_MES[mes_num],
            "FUENTE": f"FUENTE {fuente}", "COMPROBANTE": cpbte, "FECHA": fecha,
            "CLIENTE": cliente, "VENTA": round(venta, 2), "DESCUENTO": round(desc, 2),
            "VENTA NETA": venta_neta, "COSTO": round(costo, 2),
            "UTILIDAD": utilidad, "% UTILIDAD": pct_util, "VENDEDOR": vendedor,
        })
    return records


# ─── Excel I/O ────────────────────────────────────────────────────────────────

MONEY_FMT = "#,##0.00"
PCT_FMT   = "0.00%"


def _build_workbook(df: pd.DataFrame) -> Workbook:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "DATOS"
    df_out = df[COLS_DISPLAY].copy()
    HDR_COLOR  = "1F3864"
    COL_WIDTHS = {
        "AÑO":7,"MES":13,"FUENTE":11,"COMPROBANTE":14,"FECHA":10,
        "CLIENTE":36,"VENTA":16,"DESCUENTO":14,"VENTA NETA":16,
        "COSTO":16,"UTILIDAD":16,"% UTILIDAD":13,"VENDEDOR":16,
    }
    for ci, col in enumerate(COLS_DISPLAY, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=HDR_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 14)
    ws.row_dimensions[1].height = 28
    for ri, row in enumerate(df_out.itertuples(index=False), 2):
        vendedor = row[-1]
        bg   = "D6E4F0" if vendedor == _V1 else ("E2EFDA" if vendedor == _V2 else "FFFFFF")
        fill = PatternFill("solid", fgColor=bg)
        for ci, (col, value) in enumerate(zip(COLS_DISPLAY, row), 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            if col in ("VENTA","DESCUENTO","VENTA NETA","COSTO","UTILIDAD"):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col == "% UTILIDAD":
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
    last_row = len(df_out) + 1
    last_col = get_column_letter(len(COLS_DISPLAY))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}1"
    tbl = Table(displayName="VentasGilsa", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    return wb


def df_a_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    _build_workbook(df).save(buf)
    return buf.getvalue()


def leer_excel(source) -> pd.DataFrame:
    df = pd.read_excel(source, sheet_name="DATOS")
    df["MES"] = pd.Categorical(df["MES"], categories=ORDEN_MESES, ordered=True)
    df["CLIENTE_LIMPIO"] = df["CLIENTE"].apply(
        lambda x: re.sub(r"\s+\d+$", "", str(x)).strip()
    )
    if "MES_NUM" not in df.columns:
        df["MES_NUM"] = df["MES"].map(MES_A_NUM).fillna(0).astype(int)
    return df


# ─── Session state: datos ─────────────────────────────────────────────────────

def _init_data():
    if "df_data" not in st.session_state:
        if DATA_FILE.exists():
            try:
                st.session_state["df_data"] = leer_excel(DATA_FILE)
            except Exception:
                st.session_state["df_data"] = None
        else:
            st.session_state["df_data"] = None

def set_df(df: pd.DataFrame):
    df["MES"] = pd.Categorical(df["MES"], categories=ORDEN_MESES, ordered=True)
    if "CLIENTE_LIMPIO" not in df.columns:
        df["CLIENTE_LIMPIO"] = df["CLIENTE"].apply(
            lambda x: re.sub(r"\s+\d+$", "", str(x)).strip()
        )
    st.session_state["df_data"] = df


# ─── Session state: configuracion ────────────────────────────────────────────

def _init_cfg():
    if "cfg" not in st.session_state:
        loaded = {}
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, encoding="utf-8") as f:
                    loaded = json.load(f)
            except Exception:
                loaded = {}
        st.session_state["cfg"] = {**DEFAULT_CFG, **loaded}

_init_data()
_init_cfg()


# ─── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    c = cfg()

    # Logo
    if c["logo_b64"]:
        logo_bytes = base64.b64decode(c["logo_b64"])
        st.image(logo_bytes, use_container_width=True)

    st.title(c["empresa_nombre"])
    if c["empresa_nit"]:
        st.caption(f"NIT: {c['empresa_nit']}")

    st.markdown("---")
    pagina = st.radio(
        "Navegacion",
        ["Dashboard", "Importar archivos TXT", "Configuracion"],
        label_visibility="collapsed",
    )
    st.markdown("---")

    if st.session_state["df_data"] is not None:
        st.download_button(
            label="Descargar datos (Excel)",
            data=df_a_excel_bytes(st.session_state["df_data"]),
            file_name="VENTAS_GILSA_CONSOLIDADO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("Sin datos cargados.")

    st.markdown("---")


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA: CONFIGURACION
# ══════════════════════════════════════════════════════════════════════════════

if pagina == "Configuracion":

    st.title("Configuracion")

    c = cfg()

    tab_emp, tab_vend, tab_backup = st.tabs(
        ["Empresa", "Vendedores", "Guardar / Restaurar"]
    )

    # ── Tab: Empresa ──────────────────────────────────────────────────────────
    with tab_emp:
        st.subheader("Datos de la empresa")

        col_form, col_prev = st.columns([2, 1])

        with col_form:
            nuevo_nombre = st.text_input("Nombre de la empresa", value=c["empresa_nombre"])
            nuevo_nit    = st.text_input("NIT", value=c["empresa_nit"])

            st.markdown("**Logo**")
            logo_file = st.file_uploader(
                "Sube una imagen (PNG, JPG o SVG)",
                type=["png","jpg","jpeg","svg"],
                key="logo_uploader",
            )

            col_g, col_b = st.columns(2)
            guardar_emp = col_g.button("Guardar cambios", type="primary", key="btn_emp")
            if c["logo_b64"] and col_b.button("Quitar logo", key="btn_rm_logo"):
                st.session_state["cfg"]["logo_b64"]  = ""
                st.session_state["cfg"]["logo_mime"] = ""
                st.rerun()

        with col_prev:
            st.markdown("**Vista previa**")
            if logo_file:
                st.image(logo_file, use_container_width=True)
            elif c["logo_b64"]:
                st.image(base64.b64decode(c["logo_b64"]), use_container_width=True)
            else:
                st.markdown("_(sin logo)_")
            st.markdown(f"**{nuevo_nombre}**")
            if nuevo_nit:
                st.caption(f"NIT: {nuevo_nit}")

        if guardar_emp:
            st.session_state["cfg"]["empresa_nombre"] = nuevo_nombre
            st.session_state["cfg"]["empresa_nit"]    = nuevo_nit
            if logo_file:
                raw  = logo_file.read()
                mime = logo_file.type or "image/png"
                st.session_state["cfg"]["logo_b64"]  = base64.b64encode(raw).decode()
                st.session_state["cfg"]["logo_mime"] = mime
            st.success("Cambios guardados. La barra lateral se actualizara al recargar.")
            st.rerun()

    # ── Tab: Vendedores ───────────────────────────────────────────────────────
    with tab_vend:
        st.subheader("Nombres y colores de vendedores")
        st.caption(
            "Los nombres que asignes aqui aparecen en graficas, tablas y filtros. "
            "Los datos internos no cambian."
        )

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("**Vendedor 1** _(TOCARO)_")
            n1 = st.text_input("Nombre", value=c["nombre_v1"], key="n1")
            c1 = st.color_picker("Color", value=c["color_v1"], key="c1")

        with col2:
            st.markdown("**Vendedor 2** _(JUAN GABRIEL)_")
            n2 = st.text_input("Nombre", value=c["nombre_v2"], key="n2")
            c2 = st.color_picker("Color", value=c["color_v2"], key="c2")

        with col3:
            st.markdown("**Sin asignar**")
            n3 = st.text_input("Nombre", value=c["nombre_v3"], key="n3")
            c3 = st.color_picker("Color", value=c["color_v3"], key="c3")

        if st.button("Guardar cambios de vendedores", type="primary"):
            st.session_state["cfg"].update({
                "nombre_v1": n1, "color_v1": c1,
                "nombre_v2": n2, "color_v2": c2,
                "nombre_v3": n3, "color_v3": c3,
            })
            st.success("Cambios guardados.")
            st.rerun()

    # ── Tab: Guardar / Restaurar ──────────────────────────────────────────────
    with tab_backup:
        st.subheader("Persistir la configuracion")
        st.markdown(
            "La configuracion se guarda **solo en esta sesion**. "
            "Para que persista entre sesiones, descarga el archivo `config.json` "
            "y subelo a tu repositorio de GitHub junto con `dashboard_ventas.py`."
        )

        # Exportar
        cfg_export = {k: v for k, v in cfg().items()}
        st.download_button(
            label="Descargar config.json",
            data=json.dumps(cfg_export, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name="config.json",
            mime="application/json",
        )

        st.markdown("---")

        # Importar
        st.markdown("**Restaurar desde archivo**")
        cfg_file = st.file_uploader("Sube tu config.json", type=["json"], key="cfg_uploader")
        if cfg_file:
            if st.button("Aplicar configuracion", type="primary"):
                try:
                    loaded = json.loads(cfg_file.read().decode("utf-8"))
                    st.session_state["cfg"] = {**DEFAULT_CFG, **loaded}
                    st.success("Configuracion aplicada correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al leer el archivo: {e}")

        st.markdown("---")
        st.markdown("**Restablecer valores por defecto**")
        if st.button("Restablecer todo", type="secondary"):
            st.session_state["cfg"] = dict(DEFAULT_CFG)
            st.success("Configuracion restablecida.")
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA: IMPORTAR
# ══════════════════════════════════════════════════════════════════════════════

elif pagina == "Importar archivos TXT":

    st.title("Importar datos")

    tab_txt, tab_xls = st.tabs(["Subir archivos TXT", "Cargar Excel guardado"])

    with tab_txt:
        st.markdown(
            "Sube uno o varios archivos `.txt` del sistema COBOL. "
            "La app detecta automaticamente el **año, mes y fuente** desde el encabezado."
        )

        archivos = st.file_uploader(
            "Selecciona los archivos TXT",
            type=["txt"],
            accept_multiple_files=True,
            key="uploader_txt",
        )

        if archivos:
            st.subheader("Archivos detectados")
            filas_preview = []

            for archivo in archivos:
                contenido = archivo.read().decode("latin-1", errors="replace")
                lines     = contenido.splitlines()
                meta      = detectar_metadata(lines)

                if meta:
                    filas_preview.append({
                        "Archivo": archivo.name,
                        "Anio":    meta["year"],
                        "Mes":     NUM_A_MES.get(meta["mes_num"], "?"),
                        "Fuente":  f"FUENTE {meta['fuente']}",
                        "_lines":  lines, "_meta": meta, "_ok": True,
                    })
                else:
                    filas_preview.append({
                        "Archivo": archivo.name,
                        "Anio":    "No detectado",
                        "Mes":     "No detectado",
                        "Fuente":  "No detectado",
                        "_lines":  lines, "_meta": None, "_ok": False,
                    })

            df_preview = pd.DataFrame([
                {k: v for k, v in f.items() if not k.startswith("_")}
                for f in filas_preview
            ])
            st.dataframe(df_preview, use_container_width=True, hide_index=True)

            errores = [f["Archivo"] for f in filas_preview if not f["_ok"]]
            if errores:
                st.warning(
                    f"No se pudo detectar la metadata de: **{', '.join(errores)}**. "
                    "Verifica que el archivo tenga el encabezado estandar del sistema COBOL."
                )

            validos = [f for f in filas_preview if f["_ok"]]
            if not validos:
                st.error("Ningun archivo pudo ser procesado.")
                st.stop()

            st.subheader("Opciones")
            modo = st.radio(
                "Que hacer con los datos existentes?",
                ["Agregar (sin duplicar periodos)", "Reemplazar completamente"],
                key="modo_txt",
            )

            col_btn, col_info = st.columns([1, 3])
            procesar = col_btn.button("Procesar e importar", type="primary")
            col_info.markdown(f"Se procesaran **{len(validos)}** archivo(s) valido(s).")

            if procesar:
                with st.spinner("Procesando archivos..."):
                    nuevos_records = []
                    log = []
                    for f in validos:
                        meta = f["_meta"]
                        recs = parsear_contenido(
                            f["_lines"], meta["year"], meta["mes_num"], meta["fuente"]
                        )
                        nuevos_records.extend(recs)
                        log.append(
                            f"  {f['Archivo']:30} -> {len(recs):>4} facturas "
                            f"({meta['year']} {NUM_A_MES[meta['mes_num']]} FUENTE {meta['fuente']})"
                        )

                    df_nuevo = pd.DataFrame(nuevos_records)
                    df_nuevo["MES"] = pd.Categorical(
                        df_nuevo["MES"], categories=ORDEN_MESES, ordered=True
                    )
                    df_nuevo["CLIENTE_LIMPIO"] = df_nuevo["CLIENTE"].apply(
                        lambda x: re.sub(r"\s+\d+$", "", str(x)).strip()
                    )

                    df_actual = st.session_state.get("df_data")

                    if "Agregar" in modo and df_actual is not None:
                        periodos_nuevos = set(
                            zip(df_nuevo["AÑO"], df_nuevo["MES_NUM"],
                                df_nuevo["FUENTE"].str.replace("FUENTE ", "", regex=False))
                        )
                        mask = df_actual.apply(
                            lambda r: (
                                r["AÑO"],
                                MES_A_NUM.get(str(r["MES"]), 0),
                                str(r["FUENTE"]).replace("FUENTE ", ""),
                            ) in periodos_nuevos,
                            axis=1,
                        )
                        df_final = pd.concat(
                            [df_actual[~mask], df_nuevo], ignore_index=True
                        )
                    else:
                        df_final = df_nuevo

                    df_final = df_final.sort_values(
                        ["AÑO","MES_NUM","FUENTE","COMPROBANTE"]
                    ).reset_index(drop=True)
                    set_df(df_final)

                st.success(
                    f"Importacion completada. "
                    f"**{len(df_nuevo):,}** facturas nuevas — "
                    f"**{len(df_final):,}** en total."
                )
                for linea in log:
                    st.text(linea)
                st.info(
                    "Usa **Descargar datos (Excel)** en la barra lateral para guardar una copia."
                )

    with tab_xls:
        st.markdown(
            "Sube el archivo Excel que descargaste en una sesion anterior "
            "para restaurar todos los datos."
        )

        xls_file = st.file_uploader(
            "Selecciona el Excel",
            type=["xlsx"],
            key="uploader_xls",
        )

        if xls_file:
            modo_xls = st.radio(
                "Que hacer con los datos en memoria?",
                ["Reemplazar con el Excel", "Combinar con los datos actuales"],
                key="modo_xls",
            )
            if st.button("Cargar Excel", type="primary"):
                with st.spinner("Leyendo Excel..."):
                    try:
                        df_xls = leer_excel(xls_file)
                        if "Combinar" in modo_xls and st.session_state.get("df_data") is not None:
                            df_final = (
                                pd.concat([st.session_state["df_data"], df_xls], ignore_index=True)
                                .drop_duplicates(subset=["AÑO","MES_NUM","FUENTE","COMPROBANTE"])
                                .sort_values(["AÑO","MES_NUM","FUENTE","COMPROBANTE"])
                                .reset_index(drop=True)
                            )
                            set_df(df_final)
                        else:
                            set_df(df_xls)
                        st.success(
                            f"Excel cargado. "
                            f"**{len(st.session_state['df_data']):,}** facturas en memoria."
                        )
                    except Exception as e:
                        st.error(f"Error al leer el Excel: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

else:

    df_data = st.session_state.get("df_data")

    if df_data is None or df_data.empty:
        st.warning(
            "No hay datos cargados. "
            "Ve a **Importar archivos TXT** para subir tus reportes, "
            "o carga el Excel guardado de una sesion anterior."
        )
        st.stop()

    df_full = con_nombres(df_data)
    COLORES = get_colores()
    c       = cfg()

    # ── Filtros sidebar ───────────────────────────────────────────────────────
    with st.sidebar:
        st.subheader("Filtros")

        anios = sorted(df_full["AÑO"].unique())
        sel_anio = st.multiselect("Anio", anios, default=anios)

        meses_disp = [m for m in ORDEN_MESES if m in df_full["MES"].unique()]
        sel_mes = st.multiselect("Mes", meses_disp, default=meses_disp)

        fuentes_disp = sorted(df_full["FUENTE"].unique())
        sel_fuente = st.multiselect("Fuente", fuentes_disp, default=fuentes_disp)

        # Filtro vendedor usando nombres configurados
        vendedores_raw    = sorted(df_full["VENDEDOR"].unique())
        nombre_map        = get_nombre_map()
        vendedores_labels = [nombre_map.get(v, v) for v in vendedores_raw]
        inv_nombre_map    = {v: k for k, v in nombre_map.items()}
        sel_vend_labels   = st.multiselect("Vendedor", vendedores_labels, default=vendedores_labels)
        sel_vendedor      = [inv_nombre_map.get(lbl, lbl) for lbl in sel_vend_labels]

        st.markdown("---")
        st.caption(f"{c['empresa_nombre']} · NIT {c['empresa_nit']}")

    df = df_full[
        df_full["AÑO"].isin(sel_anio) &
        df_full["MES"].isin(sel_mes) &
        df_full["FUENTE"].isin(sel_fuente) &
        df_full["VENDEDOR"].isin(sel_vendedor)
    ].copy()

    # ── KPIs ──────────────────────────────────────────────────────────────────
    st.title(f"Dashboard de Ventas — {c['empresa_nombre']}")
    st.markdown(f"Mostrando **{len(df):,}** facturas")
    st.markdown("---")

    if df.empty:
        st.warning("No hay datos para los filtros seleccionados.")
        st.stop()

    venta_neta = df["VENTA NETA"].sum()
    utilidad   = df["UTILIDAD"].sum()
    pct_margen = utilidad / venta_neta * 100 if venta_neta else 0
    descuento  = df["DESCUENTO"].sum()

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Venta Neta",        f"${venta_neta/1e6:,.1f} M")
    k2.metric("Utilidad Total",    f"${utilidad/1e6:,.1f} M")
    k3.metric("% Margen Promedio", f"{pct_margen:.2f}%")
    k4.metric("Descuentos",        f"${descuento/1e6:,.1f} M")
    st.markdown("---")

    # ── Tabs ──────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "Por Mes", "Por Vendedor", "Clientes", "Detalle Facturas",
    ])

    # ── TAB 1: POR MES ────────────────────────────────────────────────────────
    with tab1:
        col_a, col_b = st.columns(2)

        mes_vend = (
            df.groupby(["AÑO","MES","VENDEDOR_DISP"], observed=True)
            .agg(UTILIDAD=("UTILIDAD","sum"), VENTA_NETA=("VENTA NETA","sum"))
            .reset_index()
        )
        mes_vend["MES_LABEL"] = mes_vend["AÑO"].astype(str) + " " + mes_vend["MES"].astype(str)

        fig1 = px.bar(
            mes_vend, x="MES_LABEL", y="UTILIDAD", color="VENDEDOR_DISP",
            color_discrete_map=COLORES, barmode="group",
            title="Utilidad por Mes y Vendedor",
            labels={"MES_LABEL":"Mes","UTILIDAD":"Utilidad ($)","VENDEDOR_DISP":"Vendedor"},
        )
        fig1.update_layout(xaxis_tickangle=-45, legend_title="Vendedor", height=420)
        col_a.plotly_chart(fig1, use_container_width=True)

        mes_total = (
            df.groupby(["AÑO","MES"], observed=True)
            .agg(UTILIDAD=("UTILIDAD","sum"), VENTA_NETA=("VENTA NETA","sum"))
            .reset_index()
        )
        mes_total["PCT"] = mes_total["UTILIDAD"] / mes_total["VENTA_NETA"] * 100
        mes_total["MES_LABEL"] = mes_total["AÑO"].astype(str) + " " + mes_total["MES"].astype(str)

        fig2 = px.line(
            mes_total, x="MES_LABEL", y="PCT", markers=True,
            title="% Margen de Utilidad por Mes",
            labels={"MES_LABEL":"Mes","PCT":"% Margen"},
            color_discrete_sequence=["#1D4ED8"],
        )
        fig2.add_hline(y=mes_total["PCT"].mean(), line_dash="dash",
                       line_color="red", annotation_text="Promedio")
        fig2.update_layout(xaxis_tickangle=-45, height=420)
        col_b.plotly_chart(fig2, use_container_width=True)

        mes_vc = (
            df.groupby(["AÑO","MES"], observed=True)
            .agg(VENTA_NETA=("VENTA NETA","sum"), COSTO=("COSTO","sum"))
            .reset_index()
        )
        mes_vc["MES_LABEL"] = mes_vc["AÑO"].astype(str) + " " + mes_vc["MES"].astype(str)
        fig3 = go.Figure()
        fig3.add_bar(x=mes_vc["MES_LABEL"], y=mes_vc["VENTA_NETA"], name="Venta Neta", marker_color="#3B82F6")
        fig3.add_bar(x=mes_vc["MES_LABEL"], y=mes_vc["COSTO"],      name="Costo",      marker_color="#EF4444")
        fig3.update_layout(barmode="group", title="Venta Neta vs Costo por Mes",
                           xaxis_tickangle=-45, height=380)
        st.plotly_chart(fig3, use_container_width=True)

    # ── TAB 2: POR VENDEDOR ───────────────────────────────────────────────────
    with tab2:
        vend_res = (
            df.groupby("VENDEDOR_DISP")
            .agg(FACTURAS=("COMPROBANTE","count"), VENTA_NETA=("VENTA NETA","sum"),
                 COSTO=("COSTO","sum"), UTILIDAD=("UTILIDAD","sum"))
            .reset_index()
        )
        vend_res["PCT_MARGEN"] = (vend_res["UTILIDAD"] / vend_res["VENTA_NETA"] * 100).round(2)

        col1, col2, col3 = st.columns(3)

        fig_pie = px.pie(vend_res, names="VENDEDOR_DISP", values="UTILIDAD",
                         color="VENDEDOR_DISP", color_discrete_map=COLORES,
                         title="Participacion en Utilidad", hole=0.45)
        col1.plotly_chart(fig_pie, use_container_width=True)

        fig_pie2 = px.pie(vend_res, names="VENDEDOR_DISP", values="VENTA_NETA",
                          color="VENDEDOR_DISP", color_discrete_map=COLORES,
                          title="Participacion en Venta Neta", hole=0.45)
        col2.plotly_chart(fig_pie2, use_container_width=True)

        fig_marg = px.bar(vend_res, x="VENDEDOR_DISP", y="PCT_MARGEN",
                          color="VENDEDOR_DISP", color_discrete_map=COLORES,
                          title="% Margen por Vendedor",
                          text=vend_res["PCT_MARGEN"].map(lambda x: f"{x:.1f}%"),
                          labels={"PCT_MARGEN":"% Margen","VENDEDOR_DISP":"Vendedor"})
        fig_marg.update_traces(textposition="outside")
        fig_marg.update_layout(showlegend=False, height=380)
        col3.plotly_chart(fig_marg, use_container_width=True)

        st.subheader("Tabla resumen por vendedor")
        tbl = vend_res.copy()
        tbl["VENTA_NETA"]  = tbl["VENTA_NETA"].map(lambda x: f"${x:,.0f}")
        tbl["COSTO"]       = tbl["COSTO"].map(lambda x: f"${x:,.0f}")
        tbl["UTILIDAD"]    = tbl["UTILIDAD"].map(lambda x: f"${x:,.0f}")
        tbl["PCT_MARGEN"]  = tbl["PCT_MARGEN"].map(lambda x: f"{x:.2f}%")
        tbl.columns = ["Vendedor","# Facturas","Venta Neta","Costo","Utilidad","% Margen"]
        st.dataframe(tbl, use_container_width=True, hide_index=True)

        st.subheader("Evolucion mensual por vendedor")
        evol = (
            df.groupby(["AÑO","MES","VENDEDOR_DISP"], observed=True)
            .agg(UTILIDAD=("UTILIDAD","sum"), VENTA_NETA=("VENTA NETA","sum"))
            .reset_index()
        )
        evol["MES_LABEL"] = evol["AÑO"].astype(str) + " " + evol["MES"].astype(str)
        evol["PCT"] = evol["UTILIDAD"] / evol["VENTA_NETA"] * 100
        fig_evol = px.line(evol, x="MES_LABEL", y="PCT", color="VENDEDOR_DISP",
                           color_discrete_map=COLORES, markers=True,
                           title="% Margen mensual por Vendedor",
                           labels={"MES_LABEL":"Mes","PCT":"% Margen","VENDEDOR_DISP":"Vendedor"})
        fig_evol.update_layout(xaxis_tickangle=-45, height=400)
        st.plotly_chart(fig_evol, use_container_width=True)

    # ── TAB 3: CLIENTES ───────────────────────────────────────────────────────
    with tab3:
        c1, c2 = st.columns([1, 3])
        with c1:
            n_top        = st.slider("Top N clientes", 5, 30, 15)
            metrica      = st.radio("Ordenar por", ["Utilidad","Venta Neta","% Margen"])
            vend_opts    = ["TODOS"] + sorted(df["VENDEDOR_DISP"].unique())
            vendedor_cli = st.selectbox("Vendedor", vend_opts)

        df_cli = df.copy() if vendedor_cli == "TODOS" else df[df["VENDEDOR_DISP"] == vendedor_cli].copy()

        cli_grp = (
            df_cli.groupby("CLIENTE_LIMPIO")
            .agg(FACTURAS=("COMPROBANTE","count"), VENTA_NETA=("VENTA NETA","sum"),
                 COSTO=("COSTO","sum"), UTILIDAD=("UTILIDAD","sum"))
            .reset_index()
        )
        cli_grp["PCT_MARGEN"] = (cli_grp["UTILIDAD"] / cli_grp["VENTA_NETA"] * 100).round(2)
        cli_grp = cli_grp[cli_grp["UTILIDAD"] > 0]

        col_ord = {"Utilidad":"UTILIDAD","Venta Neta":"VENTA_NETA","% Margen":"PCT_MARGEN"}
        top_cli = cli_grp.sort_values(col_ord[metrica], ascending=False).head(n_top)

        with c2:
            fig_cli = px.bar(
                top_cli.sort_values(col_ord[metrica]),
                x=col_ord[metrica], y="CLIENTE_LIMPIO", orientation="h",
                color="PCT_MARGEN", color_continuous_scale="Blues",
                title=f"Top {n_top} clientes — {metrica}",
                labels={"CLIENTE_LIMPIO":"","PCT_MARGEN":"% Margen"},
            )
            fig_cli.update_layout(height=max(400, n_top * 28))
            st.plotly_chart(fig_cli, use_container_width=True)

        st.subheader(f"Clientes con menor % margen — {vendedor_cli}")
        menor = cli_grp.sort_values("PCT_MARGEN").head(20)
        menor_show = menor[["CLIENTE_LIMPIO","FACTURAS","VENTA_NETA","UTILIDAD","PCT_MARGEN"]].copy()
        menor_show["VENTA_NETA"] = menor_show["VENTA_NETA"].map(lambda x: f"${x:,.0f}")
        menor_show["UTILIDAD"]   = menor_show["UTILIDAD"].map(lambda x: f"${x:,.0f}")
        menor_show["PCT_MARGEN"] = menor_show["PCT_MARGEN"].map(lambda x: f"{x:.2f}%")
        menor_show.columns = ["Cliente","Facturas","Venta Neta","Utilidad","% Margen"]
        st.dataframe(menor_show, use_container_width=True, hide_index=True)

    # ── TAB 4: DETALLE ────────────────────────────────────────────────────────
    with tab4:
        st.subheader("Detalle de facturas")
        buscar = st.text_input("Buscar cliente", placeholder="Escribe parte del nombre...")
        df_det = df.copy()
        if buscar:
            df_det = df_det[df_det["CLIENTE"].str.contains(buscar, case=False, na=False)]

        cols_mostrar = [
            "AÑO","MES","FUENTE","COMPROBANTE","FECHA","CLIENTE",
            "VENTA","DESCUENTO","VENTA NETA","COSTO","UTILIDAD","% UTILIDAD","VENDEDOR_DISP",
        ]
        df_show = df_det[cols_mostrar].copy()
        df_show["% UTILIDAD"] = (df_show["% UTILIDAD"] * 100).round(2).astype(str) + "%"
        df_show = df_show.rename(columns={"VENDEDOR_DISP": "VENDEDOR"})

        st.dataframe(
            df_show, use_container_width=True, hide_index=True, height=500,
            column_config={
                "VENTA":      st.column_config.NumberColumn(format="$%,.0f"),
                "DESCUENTO":  st.column_config.NumberColumn(format="$%,.0f"),
                "VENTA NETA": st.column_config.NumberColumn(format="$%,.0f"),
                "COSTO":      st.column_config.NumberColumn(format="$%,.0f"),
                "UTILIDAD":   st.column_config.NumberColumn(format="$%,.0f"),
            },
        )
        st.caption(f"Total filas: {len(df_det):,}")
